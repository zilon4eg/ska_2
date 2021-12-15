import access
import datetime
import os
import openpyxl
from openpyxl.styles import PatternFill
from win32com.client import Dispatch


class Excel:
    def __init__(self):
        """объявляем экземпляр класса"""
        # заносим в переменную перечень используемых цветов заливки
        self.color = {'green': r'FFCCFFCC', 'red': r'FFFF9999'}
        # заносим в переменные путь к каталогу права на папки которого нужно получить и путь к файлу excel
        path = get_all_path()
        self.file_path = path['file']
        self.dir_path = path['dir']
        # загружаем или создаем файл
        if os.path.exists(self.file_path):
            self.wb = openpyxl.open(self.file_path)
            self.ws_now = self.wb.create_sheet(self.ws_now_name(), 0)  # insert at first position
        else:
            self.wb = openpyxl.Workbook()
            self.ws_now = self.wb.active
            self.ws_now.title = self.ws_now_name()

    def ws_now_name(self):
        """создает название листа на основании текущей даты (если такой лист уже есть, добавляет время)"""
        ws_name1 = datetime.datetime.today().strftime('%d-%m-%Y')
        ws_name2 = datetime.datetime.today().strftime('%d-%m-%Y %H-%M-%S')
        return ws_name1 if ws_name1 not in self.wb.sheetnames else ws_name2

    def wb_save(self):
        """снимаем выделение со всех листов книги и сохраняем файл"""
        for sheet in self.wb:
            sheet.sheet_view.tabSelected = False
        self.wb.save(self.file_path)

    def write(self, data, sheet_number):
        """записываем данные на лист"""
        if data:
            ws = self.wb[self.wb.sheetnames[sheet_number]]
            for num_string, string in enumerate(data, 1):
                for num_cell, cell in enumerate(string, 1):
                    ws[f'{Excel.number_to_letter(num_cell)}{num_string}'].value = cell

    def get_ws_data(self, sheet_number):
        """получаем данные с листа без сервисных строк"""
        ws = self.wb[self.wb.sheetnames[sheet_number]]
        acc = []
        for string in ws.values:
            if string[0] is not None:
                acc.append(list(string))
            else:
                break
        return acc

    def add_string(self, string, num_str):
        """добавляем строку"""
        if string:
            for column, cell in enumerate(string, 1):
                self.ws_now[f'{Excel.number_to_letter(column)}{num_str}'] = cell

    def coloring_string(self, num_str, size_string, color):
        """окрашиваем строку"""
        if color is None:
            fill = PatternFill(fill_type=None)
        else:
            fill = PatternFill(start_color=self.color[color], end_color=self.color[color], fill_type='solid')
        for count in range(1, size_string + 1):
            self.ws_now[f'{Excel.number_to_letter(count)}{num_str}'].fill = fill

    @staticmethod
    def del_none(data):
        """удаляет все значения None из списка"""
        if data:
            if isinstance(data[0], list):
                return list(list(point for point in item if point is not None) for item in data)
            else:
                return list(item for item in data if item is not None)
        else:
            return data

    def coloring_on_difference(self, data_ws_0, data_ws_1):
        """окрашиваем в зеленый цвет все добавившиеся данные, все удаленные записываем сервисными строками"""
        # убираем все Null значения из данных листов
        data_ws_0 = self.del_none(data_ws_0)
        data_ws_1 = self.del_none(data_ws_1)

        # проверка на наличие данных в обоих листах
        if data_ws_0 and data_ws_1:

            if data_ws_0 != data_ws_1:
                self.coloring_tab()

            [self.coloring_string(num_string, len(str_data_ws_0), 'green') for num_string, str_data_ws_0 in enumerate(data_ws_0, 1)]

            num_redline = len(data_ws_0) + 4
            for num_string, str_data_ws_0 in enumerate(data_ws_0, 1):
                for str_data_ws_1 in data_ws_1:
                    if str_data_ws_0[0] == str_data_ws_1[0] and str_data_ws_0[1] == str_data_ws_1[1]:
                        self.coloring_string(num_string, len(str_data_ws_1), None)
                        if len(str_data_ws_0) < len(str_data_ws_1):
                            self.add_string(str_data_ws_1, num_redline)
                            self.coloring_string(num_redline, len(str_data_ws_1), 'red')
                            self.coloring_string(num_redline, len(str_data_ws_0), None)
                            num_redline += 1

            header_data_ws_0 = [[str_data_ws_0[0], str_data_ws_0[1]] for str_data_ws_0 in data_ws_0]
            for str_data_ws_1 in data_ws_1:
                if [str_data_ws_1[0], str_data_ws_1[1]] not in header_data_ws_0:
                    self.add_string(str_data_ws_1, num_redline)
                    self.coloring_string(num_redline, len(str_data_ws_1), 'red')
                    num_redline += 1

    def coloring_tab(self):
        """окрашиваем лист"""
        wsprops = self.ws_now.sheet_properties
        wsprops.tabColor = "FFFFCC00"

    @staticmethod
    def number_to_letter(number):
        """преобразует число в букву"""
        return 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'[number - 1]

    @staticmethod
    def difference_lists(list0, list1):
        """вычитает один список списков из другого"""
        set0 = set(map(lambda x: tuple(x), list0))
        set1 = set(map(lambda x: tuple(x), list1))
        set_result = set0 - set1
        list_result = list(map(lambda x: list(x), set_result))
        return list_result


def get_all_path():
    """читает пути для работы из файла"""
    if os.path.exists(f'{os.getcwd()}\\config.ini'):
        all_path = {}
        with open(f'{os.getcwd()}\\config.ini', 'r', encoding='cp1251') as file:
            for id, string in enumerate(file):
                if id == 1:
                    all_path['dir'] = string.strip().lower()
                elif id == 3:
                    all_path['file'] = string.strip().lower()
        return all_path
    else:
        return {'file': f'{os.getcwd()}\\access_list.xlsx', 'dir': f'{os.getcwd()}'}


def auto_size_column(book_path, sheet_name):
    """выбирает оптимальную ширину столбцов листа"""
    # from win32com.client import Dispatch
    excel = Dispatch('Excel.Application')
    wb = excel.Workbooks.Open(book_path)
    excel.Worksheets(sheet_name).Activate()
    excel.ActiveSheet.Columns.AutoFit()
    wb.Save()
    wb.Close()
    excel.Quit()


if __name__ == '__main__':
    # создаем экземпляр класса для работы с файлом
    xxl = Excel()
    # получаем список актуальных разрешений для папок в каталоге
    data_ws_now = access.directories_access(xxl.dir_path)
    # заносим разрешения на новый лист книги
    xxl.write(data_ws_now, 0)
    # если в книге более одного листа
    if len(xxl.wb.sheetnames) > 1:
        # получаем прошлый список разрешений (без сервисных строк) со второго листа книги
        data_ws_previous = xxl.get_ws_data(1)
        # сравниваем два листа, изменения окрашиваем, если изменения есть, окрашиваем лист.
        xxl.coloring_on_difference(data_ws_now, data_ws_previous)
    # сохраняем файл
    xxl.wb_save()
    # выставляем оптимальное значение ширины стролбцов
    auto_size_column(xxl.file_path, xxl.ws_now.title)
