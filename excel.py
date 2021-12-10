import PySimpleGUI as sg
import openpyxl
import datetime
import os
from pprint import pprint
from openpyxl.styles import PatternFill
import access
import openpyxl
from openpyxl.utils import get_column_letter
from win32com.client import Dispatch


def create_xlsx(book_path, sheet_name):
    # if not os.path.exists(book_path):
    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    sheet.title = sheet_name
    wb.save(book_path)


def edit_sheet_xlsx(book_path, sheet_name):
    if not os.path.exists(book_path):
        create_xlsx(book_path, sheet_name)
        return sheet_name
    else:
        add_sheet_name = sheet_name
        wb = openpyxl.load_workbook(book_path)
        if add_sheet_name in wb.sheetnames:
            add_sheet_name = datetime.datetime.now().strftime('%d-%m-%Y %H-%M-%S')
        wb.create_sheet(add_sheet_name)
        wb.save(book_path)
        return add_sheet_name


def insert_string(folder, user_access, book_path, sheet_name, cell_number):
    user = user_access.keys()
    user = list(user for user in user_access.keys())[0]
    add_in_cell(folder, book_path, sheet_name, cell_number, get_column_letter(1))
    add_in_cell(user, book_path, sheet_name, cell_number, get_column_letter(2))
    for id, item in enumerate(user_access[user]):
        add_in_cell(item, book_path, sheet_name, cell_number, get_column_letter(id + 3))


def add_in_cell(data, book_path, sheet_name, cell_number, cell_name):
    wb = openpyxl.load_workbook(book_path)
    sheet = wb[sheet_name]
    cell = sheet[str(cell_name) + str(cell_number)]
    cell.value = data
    wb.save(book_path)


def data_to_xls(access_list, book_path, sheet_name):
    wb = openpyxl.load_workbook(book_path)
    sheet = wb[sheet_name]
    str_num = 1
    for folder in access_list:
        if access_list[folder]:
            for user in access_list[folder]:
                insert_string(folder, user, book_path, sheet_name, str_num)
                str_num += 1


def auto_size_column(book_path, sheet_name):
    # from win32com.client import Dispatch
    excel = Dispatch('Excel.Application')
    wb = excel.Workbooks.Open(book_path)
    excel.Worksheets(sheet_name).Activate()
    excel.ActiveSheet.Columns.AutoFit()
    wb.Save()
    wb.Close()
    excel.Quit()


def color_to_string(ws, string_number, color):
    if color == 'green':
        redFill = PatternFill(start_color='FFCCFFCC', end_color='FFCCFFCC', fill_type='solid')
    else:
        redFill = PatternFill(start_color='FFFF9999', end_color='FFFF9999', fill_type='solid')
    for count in range(1, 8):
        cell_num = f'{get_column_letter(count)}{string_number}'
        ws[cell_num].fill = redFill


def sheet_to_list(book_path, sheet_name):
    wb = openpyxl.load_workbook(book_path)
    sheet = []
    ws = wb[sheet_name]
    for count in range(1, len(ws['A']) + 1):
        cells = ws[count]
        xls_str = list(cell.value for cell in cells)
        if xls_str == [None, None, None, None, None, None, None]:
            break
        sheet.append(tuple(xls_str))
    return sheet


def comparison_sheets(book_path):
    wb = openpyxl.load_workbook(book_path)
    sheetnames = wb.sheetnames
    ws = wb[sheetnames[-1]]
    if len(sheetnames) > 1:
        sheet1 = sheet_to_list(book_path, sheetnames[-1])
        sheet2 = sheet_to_list(book_path, sheetnames[-2])
        difference1 = set(sheet1) - set(sheet2)
        if difference1:
            for diff in difference1:
                color_to_string(ws, list(sheet1).index(diff) + 1, 'green')
        difference2 = set(sheet2) - set(sheet1)
        if difference2:
            report_string = len(ws['A']) + 4
            for diff in difference2:
                for id, data in enumerate(list(diff)):
                    cell = ws[str(f'{get_column_letter(id + 1)}{report_string}')]
                    cell.value = data
                color_to_string(ws, report_string, 'red')
                report_string += 1
        wb.save(book_path)


if __name__ == '__main__':
    date_today = datetime.date.today().strftime('%d-%m-%Y')
    path = r'\\fs\SHARE\Documents\OTDEL-IT\access_list.xlsx'
    root = r'\\fs\SHARE\Documents'
    xls_sheet_name = edit_sheet_xlsx(path, date_today)
    data_access = access.directories_access(root)
    data_to_xls(data_access, path, xls_sheet_name)
    auto_size_column(path, xls_sheet_name)
    comparison_sheets(path)
    print('Complete...')


    # layout = [
    #     [sg.Text('Check access'), sg.InputText(), sg.FileBrowse()],
    #     [sg.Text('File for access'), sg.InputText(), sg.FileBrowse()],
    #     [sg.Output(size=(88, 20))],
    #     [sg.Submit(), sg.Cancel()]
    # ]
    # window = sg.Window('File Compare', layout)
    # while True:                             # The Event Loop
    #     event, values = window.read()
    #     # print(event, values) #debug
    #
    #     if event in (None, 'Exit', 'Cancel'):
    #         break
    #     if event == 'Submit':
    #         date_today = datetime.date.today().strftime('%d-%m-%Y')
    #         path = r'C:\Users\suhorukov.iv\Desktop\test.xlsx'
    #         root = r'\\fs\SHARE\Documents'
    #         xls_sheet_name = edit_sheet_xlsx(path, date_today)
    #         data_access = access.directories_access(root)
    #         data_to_xls(data_access, path, xls_sheet_name)
    #         auto_size_column(path, xls_sheet_name)
    #         comparison_sheets(path)
    #         print('Complete...')
